from openpyxl import load_workbook
import json
from pathlib import Path
import numpy as np
import csv

def calculate_average(values, start_idx, end_idx):
    """
    Calculate average of numeric values in a range, ignoring 'N/A' values.
    
    Args:
        values (list): List of values
        start_idx (int): Start index (inclusive)
        end_idx (int): End index (exclusive)
        
    Returns:
        float: Average of numeric values, or 0 if no valid values
    """
    numeric_values = []
    for val in values[start_idx:end_idx]:
        if isinstance(val, (int, float)) and not np.isnan(val):
            numeric_values.append(val)
    return np.mean(numeric_values) if numeric_values else 0

def process_row_values(values, row_idx):
    """
    Process a row's values, calculating formulas instead of keeping them as strings.
    
    Args:
        values (list): List of values from the row
        row_idx (int): Row index (1-based) for reference
        
    Returns:
        list: Processed values with formulas calculated
    """
    processed_values = values.copy()
    
    # Calculate Average FLR 20y (column N)
    # Excel formula: =AVERAGE(D3:M3)
    processed_values[10] = calculate_average(values, 0, 11)  # D through M (indices 0-10)
    
    # Calculate Average plantation (column O)
    # Excel formula: =AVERAGE(D3:I3)
    processed_values[11] = calculate_average(values, 0, 6)  # D through I (indices 0-5)
    
    # Calculate Average mangrove (column P)
    # Excel formula: =IF(ISNUMBER(AVERAGE(K3:L3)),AVERAGE(K3:L3),0)
    # K and L are columns 7 and 8 in our values list (since we start from D)
    mangrove_values = [v for v in values[7:9] if isinstance(v, (int, float)) and not np.isnan(v)]
    processed_values[12] = np.mean(mangrove_values) if len(mangrove_values) == 2 else 0
    
    return processed_values

def preprocess_Winrock_data():
    """Preprocess Winrock data from Excel to JSON format"""
    data_dir = Path(__file__).parent
    
    # Load Excel workbook
    wb = load_workbook(data_dir / 'Winrock_FLR-Climate-Impact-Tool_FINAL-updated.xlsx', data_only=True)
    sheet = wb['data']
    
    # Get column headings from second row (columns D through R)
    column_headings = [sheet.cell(row=2, column=i).value for i in range(4, 19)]
    
    # Process data
    data = {}
    for row in sheet.iter_rows(min_row=3):  # Skip header rows
        country = row[1].value  # Column B
        subnational_unit = row[2].value  # Column C
        if country:
            # Get raw values from columns D through R (indices 3-17)
            raw_values = [row[i].value for i in range(3, 18)]
            
            # Process values and calculate formulas
            processed_values = process_row_values(raw_values, row[0].row)
            
            # Create nested structure: data[country][subnational_unit][column_heading] = value
            if country not in data:
                data[country] = {}
            data[country][subnational_unit] = dict(zip(column_headings, processed_values))
    
    # Save to JSON
    with open(data_dir / 'Winrock_data.json', 'w') as f:
        json.dump(data, f, indent=2)
    
    # Save to TSV for verification
    with open(data_dir / 'Winrock_data.tsv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        # Write header
        writer.writerow(['Country', 'Subnational Unit'] + column_headings)
        # Write data
        for country, units in data.items():
            for unit, values in units.items():
                writer.writerow([country, unit] + [values[heading] for heading in column_headings])
    
    print("Data preprocessing complete. Results saved to Winrock_data.json")

if __name__ == "__main__":
    preprocess_Winrock_data()